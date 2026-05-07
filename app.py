import streamlit as st
import pdfplumber
import re
import io
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sales Register – PDF to Excel",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Google Font ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* ── Hide default Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }

/* ── Hero banner ── */
.hero {
    background: linear-gradient(135deg, #1B3A6B 0%, #1d4ed8 60%, #0EA5E9 100%);
    border-radius: 16px;
    padding: 2.5rem 2rem 2rem;
    margin-bottom: 1.5rem;
    color: white;
    text-align: center;
}
.hero h1 { font-size: 2rem; font-weight: 800; margin: 0 0 .5rem; }
.hero p  { font-size: 1rem; opacity: .88; margin: 0; }

/* ── Rule cards ── */
.rules-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-bottom: 1.2rem;
}
@media(max-width:900px){ .rules-grid { grid-template-columns: repeat(2,1fr); } }
.rule-card {
    background: linear-gradient(135deg,#EFF6FF,#DBEAFE);
    border-left: 4px solid #2563EB;
    border-radius: 10px;
    padding: .85rem 1rem;
}
.rule-card.red   { background: linear-gradient(135deg,#FFF5F5,#FEE2E2); border-color:#EF4444; }
.rule-card.green { background: linear-gradient(135deg,#F0FDF4,#D1FAE5); border-color:#16A34A; }
.rule-card h4 { font-size:.8rem; font-weight:700; margin:0 0 .2rem; color:#1B3A6B; }
.rule-card p  { font-size:.75rem; color:#64748B; margin:0; line-height:1.4; }

/* ── Stat cards ── */
.stat-row { display:flex; gap:14px; margin-bottom:1.3rem; flex-wrap:wrap; }
.stat-card {
    flex:1; min-width:140px;
    background:#fff;
    border-radius:14px;
    padding:1rem 1.1rem;
    box-shadow:0 2px 12px #0001;
    border-top:3px solid #2563EB;
}
.stat-card.g { border-color:#22C55E; }
.stat-card.o { border-color:#F59E0B; }
.stat-card.v { border-color:#8B5CF6; }
.stat-label { font-size:.68rem; color:#64748B; text-transform:uppercase; letter-spacing:.5px; margin-bottom:.3rem; }
.stat-value { font-size:1.2rem; font-weight:800; color:#1B3A6B; }

/* ── Table styling ── */
.stDataFrame { border-radius:12px !important; overflow:hidden !important; }

/* ── Section headers ── */
.sec-hdr {
    font-size:.95rem; font-weight:700; color:#1B3A6B;
    border-left:4px solid #2563EB;
    padding-left:.7rem; margin:1.2rem 0 .8rem;
    display:flex; align-items:center; gap:.4rem;
}

/* ── Download button ── */
.stDownloadButton button {
    background: linear-gradient(135deg,#16A34A,#22C55E) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: .95rem !important;
    padding: .65rem 1.8rem !important;
    transition: .2s !important;
    width: 100%;
}
.stDownloadButton button:hover { transform: translateY(-1px); box-shadow:0 6px 18px #16a34a44 !important; }

/* ── Upload area override ── */
[data-testid="stFileUploader"] {
    border: 2.5px dashed #CBD5E1 !important;
    border-radius: 14px !important;
    background: #FAFCFF !important;
    padding: 1rem !important;
}

/* ── Success / Error ── */
.custom-success {
    background:#D1FAE5; border:1px solid #A7F3D0; color:#065F46;
    border-radius:10px; padding:.85rem 1.1rem; font-size:.87rem;
    margin-bottom:1rem; display:flex; align-items:center; gap:.6rem;
}
.custom-error {
    background:#FEE2E2; border:1px solid #FECACA; color:#991B1B;
    border-radius:10px; padding:.85rem 1.1rem; font-size:.87rem;
    margin-bottom:1rem;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────
MONTH_ORDER = ['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar']

# Excel colour palette
DEEP  = '1B3A6B'
MID   = '2563EB'
LIGHT = 'DBEAFE'
ALT   = 'F0F7FF'
TOTBG = 'FEF3C7'
TOTFN = '92400E'
CRBG  = 'FEE2E2'
WHITE = 'FFFFFF'


# ─────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────
def parse_pdf(pdf_bytes: bytes):
    """
    Parse SAP B1 Sales Register PDF — handles all three column formats.

    THREE FORMATS SAP B1 can produce
    ─────────────────────────────────────────────────────────────────────
    Format A  –  SGST + CGST (intra-state), 14 decimal numbers per row:
      [0]Qty [1]Price [2]Disc% [3]DiscAmt [4]SalesVal
      [5]SGST_Rate  [6]SGST_Amt
      [7]CGST_Rate  [8]CGST_Amt
      [9]IGST_Rate(0) [10]IGST_Amt(0)
      [11]TCS [12]Round [13]InvoiceTotal

    Format B  –  IGST only, SAP prints ALL columns (14 numbers), SGST/CGST = 0:
      [4]SalesVal [5]SGST_Rate(0) [6]SGST_Amt(0) [7]CGST_Rate(0) [8]CGST_Amt(0)
      [9]IGST_Rate [10]IGST_Amt [11]TCS [12]Round [13]InvoiceTotal

    Format C  –  IGST only, SAP OMITS the 4 zero SGST/CGST columns (10 numbers):
      [4]SalesVal [5]IGST_Rate [6]IGST_Amt [7]TCS [8]Round [9]InvoiceTotal

    Extraction rules
    ─────────────────
    • Sales Val  →  always floats[4]
    • SGST Amt   →  floats[6]  (Formats A & B)
    • CGST Amt   →  floats[8]  (Formats A & B)
    • IGST Amt   →  floats[10] (Formats A & B)  OR  floats[6] (Format C)
    • Format C is detected when len(floats) < 14 AND abs(floats[5]) is a known IGST rate

    Sign handling
    ─────────────
    SAP already prints negative values for credit notes, so signs are preserved
    as-is from the PDF — no manual sign flipping needed.

    Doc-type rules
    ──────────────
    • Sale Inv / Service Inv  →  positive  (added to totals)
    • Credit Note / Debit Note →  negative  (subtracted from totals)
    • Grouped by Invoice Date (first column on every data row)
    """
    monthly = defaultdict(lambda: {
        'sales_val': 0.0, 'sgst': 0.0, 'cgst': 0.0, 'igst': 0.0,
        'invoices': 0, 'credit_notes': 0,
        'month_abbr': '', 'fy_year': 0, 'dt': None,
    })
    details = []
    skipped = []

    SKIP_KW = [
        'Printed by SAP', 'Page ', 'Sales Register From', 'Invoice',
        'Customer', 'SGST', 'CGST', 'IGST', 'DETERGEO', 'PLOT NO',
        '[M]:', '[E]:', 'District', 'TAMIL NADU', 'Rate Amount',
        'dor Name', 'Round', 'TCS',
    ]

    TX = re.compile(
        r'^(\d{2}-\d{2}-\d{2,4})\s+(\d{7})\s+(CREDIT|Sale|SEVICE|Debit)\b',
        re.IGNORECASE
    )
    # Fallback: doc column is blank — line goes straight from doc number to quantity (a digit)
    TX_BLANK = re.compile(
        r'^(\d{2}-\d{2}-\d{2,4})\s+(\d{7})\s+[-\d]'
    )

    # Known IGST full rates (inter-state)
    IGST_RATES = {0.1, 0.25, 1.5, 3.0, 5.0, 6.0, 9.0, 12.0, 18.0, 28.0}
    # Known SGST/CGST half-rates (intra-state, half of the IGST rate)
    HALF_RATES = {0.05, 0.125, 0.75, 1.5, 2.5, 3.0, 4.5, 6.0, 9.0, 14.0}
    # Rates that are unambiguously IGST (never appear as SGST/CGST)
    PURE_IGST   = {5.0, 12.0, 18.0, 28.0, 3.0, 0.1, 0.25}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for raw_line in text.split('\n'):
                line = raw_line.strip()
                if not line or any(k in line for k in SKIP_KW):
                    continue

                m = TX.match(line)
                if m:
                    date_str  = m.group(1)
                    doc_no    = m.group(2)
                    dtype_raw = m.group(3).upper()
                    doc_type  = 'credit' if dtype_raw in ('CREDIT', 'DEBIT') else 'sale'
                else:
                    # Try blank-doc fallback: DATE  DOCNO  <digit/minus — no doc type word>
                    mb = TX_BLANK.match(line)
                    if not mb:
                        continue
                    date_str = mb.group(1)
                    doc_no   = mb.group(2)
                    doc_type = 'sale'   # blank doc treated as Sale Inv

                try:
                    fmt = '%d-%m-%y' if len(date_str.split('-')[2]) == 2 else '%d-%m-%Y'
                    dt  = datetime.strptime(date_str, fmt)
                except ValueError:
                    continue

                # Collect all decimal numbers from this line
                floats = []
                for tok in re.findall(r'-?[\d,]+\.\d+', line):
                    try:
                        floats.append(float(tok.replace(',', '')))
                    except ValueError:
                        pass

                # Minimum: qty price disc% discamt salesval + at least rate+amt
                if len(floats) < 7:
                    skipped.append({
                        'doc_no': doc_no, 'date': date_str,
                        'reason': f'only {len(floats)} numbers',
                        'line'  : line[:120],
                    })
                    continue

                # ── Sales Val: always at index 4 ─────────────────────
                sv = floats[4]
                if sv == 0.0:
                    continue

                sg = cg = ig = 0.0

                if len(floats) >= 14:
                    # ── Formats A & B: full 14-column layout ─────────
                    # SGST Amt=[6]  CGST Amt=[8]  IGST Amt=[10]
                    sg = floats[6]
                    cg = floats[8]
                    ig = floats[10]

                else:
                    # ── Format C (or truncated): use rate as anchor ───
                    # floats[5] = first rate column after Sales Val
                    rate = abs(floats[5])

                    if rate in PURE_IGST:
                        # IGST-only: SAP omitted SGST/CGST zero columns
                        # [5]=IGST Rate  [6]=IGST Amt
                        ig = floats[6]

                    elif rate in HALF_RATES and len(floats) >= 9:
                        # SGST+CGST short: still has SGST/CGST cols, missing tail
                        # [5]=SGST Rate  [6]=SGST Amt  [7]=CGST Rate  [8]=CGST Amt
                        sg = floats[6]
                        cg = floats[8]

                    else:
                        # Unknown layout — scan for rate→amount pairs after [4]
                        rest       = floats[5:]
                        used       = set()
                        sgst_found = False
                        for i in range(len(rest) - 1):
                            if i in used:
                                continue
                            r = abs(rest[i])
                            a = rest[i + 1]
                            if r in PURE_IGST and abs(a) > 1.0:
                                ig = a
                                used |= {i, i + 1}
                            elif r in HALF_RATES and abs(a) > 1.0:
                                if not sgst_found:
                                    sg = a
                                    sgst_found = True
                                elif cg == 0.0:
                                    cg = a
                                used |= {i, i + 1}

                # ── Accumulate into monthly buckets ──────────────────
                month_key  = dt.strftime('%b-%Y')
                month_abbr = dt.strftime('%b')
                fy_year    = dt.year if dt.month >= 4 else dt.year - 1

                mo = monthly[month_key]
                mo['sales_val']    += sv
                mo['sgst']         += sg
                mo['cgst']         += cg
                mo['igst']         += ig
                mo['month_abbr']    = month_abbr
                mo['fy_year']       = fy_year
                mo['dt']            = dt
                mo['invoices']     += (0 if doc_type == 'credit' else 1)
                mo['credit_notes'] += (1 if doc_type == 'credit' else 0)

                details.append({
                    'Doc No'      : doc_no,
                    'Type'        : 'CREDIT NOTE' if doc_type == 'credit' else 'SALE INV',
                    'Invoice Date': dt.strftime('%d-%m-%Y'),
                    'Month'       : month_key,
                    'Sales Val'   : round(sv, 2),
                    'SGST'        : round(sg, 2),
                    'CGST'        : round(cg, 2),
                    'IGST'        : round(ig, 2),
                })

    return monthly, details, skipped


def sort_keys(monthly):
    def key(k):
        mo = monthly[k]
        fy  = mo.get('fy_year', 0)
        idx = MONTH_ORDER.index(mo.get('month_abbr', 'Apr')) \
              if mo.get('month_abbr') in MONTH_ORDER else 0
        return (fy, idx)
    return sorted(monthly.keys(), key=key)


# ─────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────
def _hdr(cell, bg=MID, txt=WHITE, sz=10):
    cell.font      = Font(bold=True, color=txt, size=sz, name='Arial')
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _borders(ws, rng, c='B0C4DE'):
    s = Side(style='thin', color=c)
    b = Border(left=s, right=s, top=s, bottom=s)
    for row in ws[rng]:
        for cell in row:
            cell.border = b

def _money(cell, val, fg, bold=False, neg=False):
    cell.value         = val
    cell.number_format = '#,##0.00'
    cell.font          = Font(name='Arial', size=9, bold=bold,
                               color='CC3300' if neg else '1E293B')
    cell.fill          = PatternFill('solid', fgColor=fg)
    cell.alignment     = Alignment(horizontal='right', vertical='center')

def _text(cell, val, fg, bold=False, center=False, color='1E293B', sz=9):
    cell.value     = val
    cell.font      = Font(name='Arial', size=sz, bold=bold, color=color)
    cell.fill      = PatternFill('solid', fgColor=fg)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left', vertical='center')


def build_excel(monthly, details) -> bytes:
    wb = Workbook()
    sorted_k = sort_keys(monthly)

    # ── Sheet 1 : Monthly Summary ────────────────────────────────────
    ws = wb.active
    ws.title = 'Monthly Summary'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:I1')
    ws['A1'] = '  DETERGEO CHEM PVT. LTD. — Sales Register (Monthly)'
    ws['A1'].font      = Font(bold=True, color=WHITE, size=14, name='Arial')
    ws['A1'].fill      = PatternFill('solid', fgColor=DEEP)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:I2')
    ws['A2'] = f'  Generated: {datetime.now().strftime("%d %b %Y, %H:%M")}  |  Period: 01-Apr-2025 to 31-Mar-2026'
    ws['A2'].font      = Font(color='444444', size=9, name='Arial', italic=True)
    ws['A2'].fill      = PatternFill('solid', fgColor=LIGHT)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    heads  = ['Month','Sales Value (₹)','SGST (₹)','CGST (₹)','IGST (₹)',
              'Total Tax (₹)','Gross Total (₹)','Invoices','Credit Notes']
    cwidths = [14, 20, 18, 18, 18, 18, 20, 11, 13]
    for c, (h, w) in enumerate(zip(heads, cwidths), 1):
        cell = ws.cell(4, c, h)
        _hdr(cell)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[4].height = 28

    tot = {'sv': 0.0, 'sg': 0.0, 'cg': 0.0, 'ig': 0.0}
    r = 5
    for idx, key in enumerate(sorted_k):
        m  = monthly[key]
        sv = round(m['sales_val'], 2)
        sg = round(m['sgst'], 2)
        cg = round(m['cgst'], 2)
        ig = round(m['igst'], 2)
        tx = round(sg + cg + ig, 2)
        gr = round(sv + tx, 2)
        tot['sv'] += sv; tot['sg'] += sg; tot['cg'] += cg; tot['ig'] += ig

        bg  = CRBG if sv < 0 else (ALT if idx % 2 == 0 else WHITE)
        neg = sv < 0

        _text (ws.cell(r, 1), key,             bg, center=True)
        _money(ws.cell(r, 2), sv,              bg, neg=neg)
        _money(ws.cell(r, 3), sg,              bg, neg=neg)
        _money(ws.cell(r, 4), cg,              bg, neg=neg)
        _money(ws.cell(r, 5), ig,              bg, neg=neg)
        _money(ws.cell(r, 6), tx,              bg, neg=neg)
        _money(ws.cell(r, 7), gr,              bg, bold=True, neg=neg)
        _text (ws.cell(r, 8), m['invoices'],   bg, center=True)
        _text (ws.cell(r, 9), m['credit_notes'], bg, center=True,
               color='CC3300' if m['credit_notes'] else '1E293B')
        ws.row_dimensions[r].height = 20
        r += 1

    ttx = round(tot['sg'] + tot['cg'] + tot['ig'], 2)
    tgr = round(tot['sv'] + ttx, 2)
    for c, v in enumerate(['GRAND TOTAL', tot['sv'], tot['sg'], tot['cg'],
                            tot['ig'], ttx, tgr, '', ''], 1):
        cell = ws.cell(r, c, v)
        cell.fill = PatternFill('solid', fgColor=TOTBG)
        cell.font = Font(bold=True, name='Arial', size=10, color=TOTFN)
        cell.alignment = Alignment(
            horizontal='center' if c in (1, 8, 9) else 'right', vertical='center')
        if isinstance(v, float):
            cell.number_format = '#,##0.00'
    ws.row_dimensions[r].height = 26
    _borders(ws, f'A4:I{r}')
    ws.freeze_panes = 'A5'

    # ── Sheet 2 : Transactions ───────────────────────────────────────
    ws2 = wb.create_sheet('Transactions')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:H1')
    ws2['A1'] = '  TRANSACTION DETAIL LOG'
    ws2['A1'].font      = Font(bold=True, color=WHITE, size=13, name='Arial')
    ws2['A1'].fill      = PatternFill('solid', fgColor=DEEP)
    ws2['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells('A2:H2')
    sales_cnt  = sum(1 for d in details if d['Type'] == 'SALE INV')
    credit_cnt = sum(1 for d in details if d['Type'] == 'CREDIT NOTE')
    ws2['A2'] = (f'  Total: {len(details)} records  |  '
                 f'Sale Inv: {sales_cnt}  |  Credit Notes: {credit_cnt}')
    ws2['A2'].font      = Font(color='444444', size=9, name='Arial', italic=True)
    ws2['A2'].fill      = PatternFill('solid', fgColor=LIGHT)
    ws2['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[2].height = 16
    ws2.row_dimensions[3].height = 6

    dh = ['Doc No','Type','Invoice Date','Month',
          'Sales Val (₹)','SGST (₹)','CGST (₹)','IGST (₹)']
    dw = [12, 15, 16, 12, 20, 18, 18, 18]
    for c, (h, w) in enumerate(zip(dh, dw), 1):
        cell = ws2.cell(4, c, h)
        _hdr(cell)
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[4].height = 26

    for idx, rec in enumerate(details):
        r2   = idx + 5
        iscr = rec['Type'] == 'CREDIT NOTE'
        bg   = CRBG if iscr else (ALT if idx % 2 == 0 else WHITE)
        _text (ws2.cell(r2,1), rec['Doc No'],       bg, center=True)
        _text (ws2.cell(r2,2), rec['Type'],         bg, bold=iscr, center=True,
               color='CC3300' if iscr else '1B3A6B')
        _text (ws2.cell(r2,3), rec['Invoice Date'], bg, center=True)
        _text (ws2.cell(r2,4), rec['Month'],        bg, center=True)
        _money(ws2.cell(r2,5), rec['Sales Val'],    bg, neg=iscr)
        _money(ws2.cell(r2,6), rec['SGST'],         bg, neg=iscr)
        _money(ws2.cell(r2,7), rec['CGST'],         bg, neg=iscr)
        _money(ws2.cell(r2,8), rec['IGST'],         bg, neg=iscr)
        ws2.row_dimensions[r2].height = 18

    if details:
        _borders(ws2, f'A4:H{len(details)+4}')
    ws2.freeze_panes = 'A5'

    # ── Sheet 3 : Quarter View ───────────────────────────────────────
    ws3 = wb.create_sheet('Quarter View')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:F1')
    ws3['A1'] = '  QUARTERLY ANALYSIS'
    ws3['A1'].font      = Font(bold=True, color=WHITE, size=13, name='Arial')
    ws3['A1'].fill      = PatternFill('solid', fgColor=DEEP)
    ws3['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[1].height = 30
    ws3.row_dimensions[2].height = 6

    qh = ['Quarter','Sales Value (₹)','SGST (₹)','CGST (₹)','IGST (₹)','Total Tax (₹)']
    qw = [20, 22, 18, 18, 18, 18]
    for c, (h, w) in enumerate(zip(qh, qw), 1):
        cell = ws3.cell(3, c, h)
        _hdr(cell)
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.row_dimensions[3].height = 26

    quarters = [
        ('Q1 (Apr–Jun)', ['Apr','May','Jun'], 'E0F2FE'),
        ('Q2 (Jul–Sep)', ['Jul','Aug','Sep'], 'EDE9FE'),
        ('Q3 (Oct–Dec)', ['Oct','Nov','Dec'], 'FEF3C7'),
        ('Q4 (Jan–Mar)', ['Jan','Feb','Mar'], 'FCE7F3'),
    ]
    qr = 4
    for qname, months, bg in quarters:
        sv = sg = cg = ig = 0.0
        for key in sorted_k:
            if monthly[key].get('month_abbr','') in months:
                sv += monthly[key]['sales_val']
                sg += monthly[key]['sgst']
                cg += monthly[key]['cgst']
                ig += monthly[key]['igst']
        tx = sg + cg + ig
        for c, val in enumerate([qname,round(sv,2),round(sg,2),
                                  round(cg,2),round(ig,2),round(tx,2)], 1):
            cell = ws3.cell(qr, c, val)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.font = Font(bold=True, name='Arial', size=10, color='1E293B')
            cell.alignment = Alignment(
                horizontal='center' if c==1 else 'right', vertical='center')
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
        ws3.row_dimensions[qr].height = 22
        qr += 1

    for c in range(1, 7):
        cell = ws3.cell(qr, c)
        cell.fill = PatternFill('solid', fgColor=TOTBG)
        cell.font = Font(bold=True, name='Arial', size=10, color=TOTFN)
        cell.alignment = Alignment(
            horizontal='center' if c==1 else 'right', vertical='center')
        if c == 1:
            cell.value = 'TOTAL'
        else:
            cell.value         = f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{qr-1})'
            cell.number_format = '#,##0.00'
    ws3.row_dimensions[qr].height = 24
    _borders(ws3, f'A3:F{qr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def fmt_inr(n: float) -> str:
    """Format number as INR with commas."""
    return f"₹ {n:,.2f}"


def fmt_big(n: float) -> str:
    """Compact INR: Cr / L / plain."""
    a = abs(n)
    if a >= 1e7:
        s = f"{a/1e7:.2f} Cr"
    elif a >= 1e5:
        s = f"{a/1e5:.1f} L"
    else:
        s = f"{a:,.2f}"
    return ("-" if n < 0 else "") + "₹ " + s


# ─────────────────────────────────────────────────────────────
# UI — HEADER
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <h1>📊 Sales Register PDF → Monthly Excel</h1>
  <p>Upload your SAP Business One Sales Register PDF and get a formatted,
     month-wise Excel report with Sales Value, SGST, CGST &amp; IGST —
     credit notes handled automatically.</p>
</div>
""", unsafe_allow_html=True)

# ── Processing rules ─────────────────────────────────────────
st.markdown("""
<div class="rules-grid">
  <div class="rule-card">
    <h4>📄 Sale Invoice</h4>
    <p>Treated as <strong>positive</strong> — added to monthly totals.</p>
  </div>
  <div class="rule-card green">
    <h4>🔧 Service Invoice</h4>
    <p>Also <strong>positive</strong> — added to sales totals.</p>
  </div>
  <div class="rule-card red">
    <h4>↩️ Credit Note</h4>
    <p><strong>Negative</strong> — subtracted from monthly totals.</p>
  </div>
  <div class="rule-card">
    <h4>📅 Invoice Date</h4>
    <p>Grouped by Invoice Date. FY order: Apr → Mar.</p>
  </div>
</div>
""", unsafe_allow_html=True)

st.divider()

# ─────────────────────────────────────────────────────────────
# UI — UPLOAD
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="sec-hdr">📁 Upload Sales Register PDF</div>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    label="Drop your SAP Business One Sales Register PDF here",
    type=["pdf"],
    help="Supports SAP B1 Sales Register exports · Max 50 MB",
)

if uploaded:
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.success(f"✅ **{uploaded.name}** — {uploaded.size / 1024 / 1024:.2f} MB")
    with col2:
        process_btn = st.button("⚡ Process PDF", type="primary", use_container_width=True)
    with col3:
        st.write("")   # spacer

    if process_btn:
        with st.spinner("🔍 Scanning PDF and extracting transactions…"):
            try:
                pdf_bytes = uploaded.read()
                monthly, details, skipped = parse_pdf(pdf_bytes)

                if not monthly:
                    st.markdown("""<div class="custom-error">
                        ⚠️ <strong>No transaction data found.</strong>
                        Please check the PDF is a SAP Business One Sales Register export.
                    </div>""", unsafe_allow_html=True)
                    st.stop()

                # ── Build summary rows ────────────────────────────────
                sorted_k   = sort_keys(monthly)
                total_sv   = sum(monthly[k]['sales_val'] for k in sorted_k)
                total_sg   = sum(monthly[k]['sgst']      for k in sorted_k)
                total_cg   = sum(monthly[k]['cgst']      for k in sorted_k)
                total_ig   = sum(monthly[k]['igst']      for k in sorted_k)
                total_tax  = total_sg + total_cg + total_ig
                total_inv  = sum(monthly[k]['invoices']      for k in sorted_k)
                total_cn   = sum(monthly[k]['credit_notes']  for k in sorted_k)

                # ── Stat cards ────────────────────────────────────────
                st.markdown(f"""
                <div class="stat-row">
                  <div class="stat-card">
                    <div class="stat-label">Total Sales Value</div>
                    <div class="stat-value">{fmt_big(total_sv)}</div>
                  </div>
                  <div class="stat-card g">
                    <div class="stat-label">Total Tax (GST)</div>
                    <div class="stat-value">{fmt_big(total_tax)}</div>
                  </div>
                  <div class="stat-card o">
                    <div class="stat-label">Months Found</div>
                    <div class="stat-value">{len(sorted_k)}</div>
                  </div>
                  <div class="stat-card v">
                    <div class="stat-label">Invoices / CNs</div>
                    <div class="stat-value">{total_inv} / {total_cn}</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Monthly table ─────────────────────────────────────
                st.markdown('<div class="sec-hdr">📋 Monthly Breakdown</div>',
                            unsafe_allow_html=True)

                import pandas as pd
                rows = []
                for key in sorted_k:
                    m  = monthly[key]
                    sv = round(m['sales_val'], 2)
                    sg = round(m['sgst'], 2)
                    cg = round(m['cgst'], 2)
                    ig = round(m['igst'], 2)
                    tx = round(sg + cg + ig, 2)
                    gr = round(sv + tx, 2)
                    inv_str = f"{m['invoices']} Inv"
                    if m['credit_notes']:
                        inv_str += f"  /  {m['credit_notes']} CN"
                    rows.append({
                        'Month'         : key,
                        'Sales Value'   : sv,
                        'SGST'          : sg,
                        'CGST'          : cg,
                        'IGST'          : ig,
                        'Total Tax'     : tx,
                        'Gross Total'   : gr,
                        'Inv / CN'      : inv_str,
                    })

                df = pd.DataFrame(rows)

                # Totals row
                tot_row = pd.DataFrame([{
                    'Month'      : '📊 GRAND TOTAL',
                    'Sales Value': round(total_sv, 2),
                    'SGST'       : round(total_sg, 2),
                    'CGST'       : round(total_cg, 2),
                    'IGST'       : round(total_ig, 2),
                    'Total Tax'  : round(total_tax, 2),
                    'Gross Total': round(total_sv + total_tax, 2),
                    'Inv / CN'   : '',
                }])
                df_display = pd.concat([df, tot_row], ignore_index=True)

                money_cols = ['Sales Value','SGST','CGST','IGST','Total Tax','Gross Total']

                st.dataframe(
                    df_display.style
                        .format({c: "₹ {:,.2f}" for c in money_cols})
                        .apply(lambda row: [
                            'background-color:#FEE2E2; color:#DC2626'
                            if row['Sales Value'] < 0
                            else ('background-color:#FEF9C3; font-weight:bold'
                                  if row['Month'] == '📊 GRAND TOTAL'
                                  else '')
                            for _ in row
                        ], axis=1),
                    use_container_width=True,
                    hide_index=True,
                    height=min(38 * (len(df_display) + 1) + 38, 600),
                )

                # ── Transaction detail expander ───────────────────────
                with st.expander(f"🔍 View All {len(details)} Transactions"):
                    det_df = pd.DataFrame(details)
                    money_det = ['Sales Val','SGST','CGST','IGST']
                    st.dataframe(
                        det_df.style
                            .format({c: "₹ {:,.2f}" for c in money_det})
                            .apply(lambda row: [
                                'background-color:#FEE2E2; color:#DC2626'
                                if row['Type'] == 'CREDIT NOTE' else ''
                                for _ in row
                            ], axis=1),
                        use_container_width=True,
                        hide_index=True,
                    )

                # ── Skipped lines debug ───────────────────────────────
                if skipped:
                    with st.expander(
                        f"⚠️ {len(skipped)} lines could not be parsed (click to inspect)",
                        expanded=False
                    ):
                        st.caption(
                            "These lines matched a date + doc number pattern but lacked "
                            "enough numeric columns. They are **excluded** from totals. "
                            "If any are valid transactions, the PDF format may differ — "
                            "share this list for a fix."
                        )
                        st.dataframe(
                            pd.DataFrame(skipped)[['doc_no','date','reason','line']],
                            use_container_width=True,
                            hide_index=True,
                        )

                # ── Build & offer Excel download ──────────────────────
                st.markdown('<div class="sec-hdr">⬇️ Download Excel Report</div>',
                            unsafe_allow_html=True)

                xlsx_bytes = build_excel(monthly, details)
                dl_col1, dl_col2, dl_col3 = st.columns([1, 2, 1])
                with dl_col2:
                    st.download_button(
                        label="⬇️  Download Excel Report",
                        data=xlsx_bytes,
                        file_name="Sales_Register_Monthly.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

            except Exception as e:
                import traceback
                st.error(f"**Processing Error:** {e}")
                with st.expander("Show full traceback"):
                    st.code(traceback.format_exc())

else:
    st.info("⬆️  Upload a PDF above to get started.")

# ─────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────
st.divider()
st.markdown(
    "<p style='text-align:center;color:#94A3B8;font-size:.75rem;'>"
    "Detergeo Chem Pvt. Ltd. · Sales Register Extractor · "
    "Built with Python · pdfplumber · Streamlit · openpyxl"
    "</p>",
    unsafe_allow_html=True,
)

